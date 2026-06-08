#!/usr/bin/env python3
"""
Call Break Scoreboard — Migration Script (fixed)
Reads callbreak_migration.xlsx and inserts data into your PostgreSQL/Supabase DB.

Usage:
  pip install openpyxl psycopg2-binary
  python migrate_fixed.py

Set your DB connection string in DATABASE_URL below (or as an env variable).

Fixes applied vs original:
  1. Load workbook with data_only=True so Excel formula cells (=Sheet5!A3 etc.)
     are resolved to their cached values instead of returning the raw formula string.
  2. The 'p1_actual_wins' … 'p4_actual_wins' columns store floats (e.g. 3.1, -4).
     They are now read as floats and passed as-is to the DB — no recalculation.
  3. Players sheet may be empty; players are auto-created from the Matches sheet
     p_names columns when absent, using the group_id from the match row.
  4. group_id was read from the Players sheet header but never stored in the
     players dict, causing a KeyError on upsert.  It is now stored and used.
"""

import os
import sys
from datetime import datetime
from decimal import Decimal

import openpyxl
import psycopg2
from psycopg2.extras import RealDictCursor

# ──────────────────────────────────────────────────────────────────────────────
# CONFIG — edit this or set the DATABASE_URL environment variable
# ──────────────────────────────────────────────────────────────────────────────
DATABASE_URL = os.getenv(
    "DATABASE_URL",
    "postgresql://postgres:BreakArena%407878@db.glgepcayubpuyfonndzu.supabase.co:5432/postgres"
)
EXCEL_FILE = os.getenv("EXCEL_FILE", "callbreak_migration.xlsx")

# ──────────────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────────────
def cell_val(ws, row, col):
    return ws.cell(row=row, column=col).value

def str_val(ws, row, col):
    v = cell_val(ws, row, col)
    return str(v).strip() if v is not None else None

def int_val(ws, row, col):
    v = cell_val(ws, row, col)
    if v is None:
        return None
    # Guard against formula strings that survived despite data_only=True
    if isinstance(v, str) and v.startswith("="):
        return None
    return int(float(v))

def float_val(ws, row, col):
    v = cell_val(ws, row, col)
    if v is None:
        return None
    if isinstance(v, str) and v.startswith("="):
        return None
    return float(v)


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────
def main():
    print(f"📂  Reading {EXCEL_FILE} …")
    # FIX 1: data_only=True resolves cached formula values instead of returning
    #         the raw formula string (e.g. '=Sheet5!A3').
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    # ── Read Players sheet ────────────────────────────────────────────────────
    players_ws = wb["Players"]
    players = {}  # name (lower) → dict
    for row in range(2, players_ws.max_row + 1):
        name = str_val(players_ws, row, 2)
        if not name or name.startswith("←"):
            continue
        color    = str_val(players_ws, row, 3) or "#fbbf24"
        avatar   = str_val(players_ws, row, 4) or ""
        is_bot_raw = str_val(players_ws, row, 5) or "False"
        is_bot   = is_bot_raw.strip().lower() in ("true", "1", "yes")
        # FIX 4: store group_id from the Players sheet
        group_id = str_val(players_ws, row, 6)
        players[name.lower()] = {
            "name": name, "color": color,
            "avatar": avatar, "is_bot": is_bot,
            "group_id": group_id,
            "id": None,
        }
    print(f"  → {len(players)} player(s) found in Players sheet: "
          f"{[p['name'] for p in players.values()]}")

    # ── Read Matches sheet ────────────────────────────────────────────────────
    matches_ws = wb["Matches"]
    matches = {}  # match_number → dict
    for row in range(2, matches_ws.max_row + 1):
        mnum = int_val(matches_ws, row, 1)
        if mnum is None:
            continue
        raw_date = cell_val(matches_ws, row, 2)
        if isinstance(raw_date, str):
            match_date = raw_date.strip()
        elif hasattr(raw_date, "strftime"):
            match_date = raw_date.strftime("%Y-%m-%d")
        else:
            match_date = datetime.today().strftime("%Y-%m-%d")
        total_rounds = int_val(matches_ws, row, 3) or 13
        p_names      = [str_val(matches_ws, row, c) for c in range(4, 8)]
        status       = str_val(matches_ws, row, 8) or "completed"
        group_id     = str_val(matches_ws, row, 9)
        matches[mnum] = {
            "match_number":  mnum,
            "match_date":    match_date,
            "total_rounds":  total_rounds,
            "p_names":       p_names,   # [p1, p2, p3, p4]
            "status":        status,
            "group_id":      group_id,
            "id":            None,
        }

        # FIX 3: if Players sheet was empty, auto-create player stubs from
        #         the match row so the rest of the script has something to work with.
        for pname in p_names:
            if pname and pname.lower() not in players:
                players[pname.lower()] = {
                    "name":     pname,
                    "color":    "#fbbf24",
                    "avatar":   "",
                    "is_bot":   False,
                    "group_id": group_id,
                    "id":       None,
                }

    print(f"  → {len(matches)} match(es) found")
    print(f"  → {len(players)} unique player(s) after merging with Matches: "
          f"{[p['name'] for p in players.values()]}")

    # ── Read Rounds sheet ─────────────────────────────────────────────────────
    rounds_ws = wb["Rounds"]
    all_rounds = []
    for row in range(2, rounds_ws.max_row + 1):
        mnum = int_val(rounds_ws, row, 1)
        rnum = int_val(rounds_ws, row, 2)
        if mnum is None or rnum is None:
            continue

        # FIX 2: read bids as int, but read 'actual_wins' columns as float
        #         because they store pre-computed round scores.
        bids   = [int_val(rounds_ws, row, c)   for c in (3, 5, 7, 9)]
        scores = [float_val(rounds_ws, row, c) for c in (4, 6, 8, 10)]

        if any(v is None for v in bids + scores):
            print(f"  ⚠  Skipping round row {row}: missing bid/score data")
            continue

        all_rounds.append({
            "match_number": mnum,
            "round_number": rnum,
            "bids":         bids,
            "scores":       scores,   # pre-computed round scores from sheet
        })
    print(f"  → {len(all_rounds)} round row(s) found")

    # ── Connect to DB ─────────────────────────────────────────────────────────
    print(f"\n🔌  Connecting to database …")
    try:
        conn = psycopg2.connect(DATABASE_URL)
    except Exception as e:
        print(f"  ✗ Could not connect: {e}")
        sys.exit(1)

    cur = conn.cursor(cursor_factory=RealDictCursor)

    # ── Upsert players ────────────────────────────────────────────────────────
    print("👤  Upserting players …")
    for key, p in players.items():
        cur.execute(
            """
            INSERT INTO players (name, color, avatar, is_bot, group_id)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT DO NOTHING
            RETURNING id
            """,
            (p["name"], p["color"], p["avatar"], p["is_bot"], p["group_id"])
        )
        row = cur.fetchone()
        if row:
            p["id"] = row["id"]
        else:
            cur.execute(
                "SELECT id FROM players WHERE LOWER(name) = LOWER(%s)",
                (p["name"],)
            )
            p["id"] = cur.fetchone()["id"]
        print(f"  ✓ {p['name']} → {p['id']}")

    # ── Insert matches & rounds ───────────────────────────────────────────────
    print("🃏  Inserting matches and rounds …")
    for mnum, m in matches.items():
        p_ids = []
        for pname in m["p_names"]:
            pid = players.get(pname.lower(), {}).get("id") if pname else None
            p_ids.append(pid)

        match_rounds = sorted(
            [r for r in all_rounds if r["match_number"] == mnum],
            key=lambda x: x["round_number"]
        )

        cur.execute(
            """
            INSERT INTO matches
              (match_date, match_number, total_rounds, status,
               p1_id, p2_id, p3_id, p4_id,
               p1_total_score, p2_total_score, p3_total_score, p4_total_score,
               group_id)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,0,0,0,0,%s)
            ON CONFLICT (match_date, match_number) DO UPDATE
              SET status = EXCLUDED.status
            RETURNING id
            """,
            (m["match_date"], mnum, m["total_rounds"], m["status"],
             p_ids[0], p_ids[1], p_ids[2], p_ids[3], m["group_id"])
        )
        m["id"] = cur.fetchone()["id"]
        print(f"  ✓ Match #{mnum} → {m['id']}")

        # Running totals — accumulated from stored per-round scores
        running = [Decimal("0"), Decimal("0"), Decimal("0"), Decimal("0")]

        for r in match_rounds:
            for i in range(4):
                running[i] += Decimal(str(r["scores"][i]))

            cur.execute(
                """
                INSERT INTO rounds
                  (match_id, round_number,
                   p1_bid, p1_actual_wins, p1_total_score,
                   p2_bid, p2_actual_wins, p2_total_score,
                   p3_bid, p3_actual_wins, p3_total_score,
                   p4_bid, p4_actual_wins, p4_total_score)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (match_id, round_number) DO UPDATE
                  SET p1_bid=EXCLUDED.p1_bid, p1_actual_wins=EXCLUDED.p1_actual_wins,
                      p1_total_score=EXCLUDED.p1_total_score,
                      p2_bid=EXCLUDED.p2_bid, p2_actual_wins=EXCLUDED.p2_actual_wins,
                      p2_total_score=EXCLUDED.p2_total_score,
                      p3_bid=EXCLUDED.p3_bid, p3_actual_wins=EXCLUDED.p3_actual_wins,
                      p3_total_score=EXCLUDED.p3_total_score,
                      p4_bid=EXCLUDED.p4_bid, p4_actual_wins=EXCLUDED.p4_actual_wins,
                      p4_total_score=EXCLUDED.p4_total_score
                """,
                (
                    m["id"], r["round_number"],
                    r["bids"][0], r["scores"][0], float(running[0]),
                    r["bids"][1], r["scores"][1], float(running[1]),
                    r["bids"][2], r["scores"][2], float(running[2]),
                    r["bids"][3], r["scores"][3], float(running[3]),
                )
            )
            print(f"    Round {r['round_number']:2d} | actual_wins: "
                  + " | ".join(f"{s} (Σ{float(t):.1f})"
                                for s, t in zip(r["scores"], running)))

        # Update match totals and winner
        cur.execute(
            """
            UPDATE matches SET
              p1_total_score=%s, p2_total_score=%s,
              p3_total_score=%s, p4_total_score=%s,
              ended_at = CASE WHEN status='completed'
                              THEN NOW() AT TIME ZONE 'Asia/Kolkata'
                              ELSE ended_at END
            WHERE id=%s
            """,
            (float(running[0]), float(running[1]),
             float(running[2]), float(running[3]), m["id"])
        )

        ranked = sorted(
            zip(p_ids, running),
            key=lambda x: x[1], reverse=True
        )
        winner_updates = [r[0] for r in ranked]
        cur.execute(
            """
            UPDATE matches SET
              winner_id=%s, sec_winner_id=%s,
              third_winner_id=%s, fourth_winner_id=%s
            WHERE id=%s
            """,
            (winner_updates[0], winner_updates[1],
             winner_updates[2], winner_updates[3], m["id"])
        )

    conn.commit()
    cur.close()
    conn.close()
    print("\n✅  Migration complete!")


if __name__ == "__main__":
    main()